"""export_top_videos 的确定性部分单测（不调用 VLM/网络）。

覆盖：tier 定档、复合排名、规则判定（通过/未取证/待接入/外部依赖/不合格）、
xlsx 五表结构与关键单元格。
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[2]


def _run(name: str, avg: float, minimum: float, *, cert: bool = True, l1a: str = "pass",
         checks: dict | None = None, noncut: int = 0, cuts: int = 8, cost: float = 0.053,
         temp: Path | None = None) -> dict:
    checks = checks or {}
    return {
        "run": name, "project": temp or ROOT / "projects" / name,
        "sheet_name": f"sheet-{name}", "template_id": name,
        "l1a": {"status": l1a}, "l1a_checks": checks, "qa": {}, "probe": {
            "resolution": "1080x1920", "duration_seconds": 40.0, "has_audio": True},
        "certificate": {"version": "1.0"} if cert else None, "edit": {}, "cuts": [{}] * cuts,
        "noncut": noncut, "script": {"total_duration_seconds": 40.0}, "duration_s": 40.0,
        "cost": cost, "cost_note": "", "published_at": "2026-08-26T00:00:00Z",
        "advisory": {"dimensions": {"hook_clarity": minimum + 0.5, "rhythm": avg,
                                    "story_coherence": minimum, "product_presence": avg}},
        "l3_avg": avg, "l3_min": minimum, "weakest": "story_coherence" if minimum <= 6.5 else "rhythm",
    }


def test_tier_and_rank():
    from scripts.export_top_videos import rank_runs, tier_of

    gold = _run("gold", 8.6, 8.0, cert=False)
    good = _run("good", 8.3, 7.5)
    weak = _run("weak", 7.8, 6.0, noncut=1)
    assert tier_of(gold) == "推荐"
    assert tier_of(good) == "达标"
    assert "观察" in tier_of(weak)
    order = rank_runs([weak, good, gold])
    assert [r["run"] for r in order] == ["gold", "good", "weak"]
    # 同分：有证书优先（tie-break）
    tie_a = _run("tie-no-cert", 8.3, 7.5, cert=False)
    tie_b = _run("tie-cert", 8.3, 7.5, cert=True)
    assert [r["run"] for r in rank_runs([tie_a, tie_b])][0] == "tie-cert"


def test_rule_verdict_layers():
    from scripts.export_top_videos import rule_verdict

    run = _run("x", 8.0, 7.0, checks={"l1a_sensitive": {"status": "pass"}})
    assert rule_verdict(
        {"layer": "hard_gate", "checks": ["l1a_sensitive"], "severity": "veto"}, run)[1] == "pass"
    assert rule_verdict({"layer": "external", "checks": []}, run)[1] == "external"
    assert rule_verdict({"layer": "hard_gate", "checks": []}, run)[1] == "pending"
    fail_run = _run("f", 8.0, 7.0, checks={"l1a_sensitive": {"status": "fail", "message": "命中 2"}})
    text, style = rule_verdict({"layer": "hard_gate", "checks": ["l1a_sensitive"]}, fail_run)
    assert style == "fail" and "不合格" in text


def test_workbook_structure(tmp_path: Path):
    import yaml as _yaml
    from openpyxl import load_workbook

    from scripts.export_top_videos import write_workbook

    policy = _yaml.safe_load((ROOT / "docs/rules/business-policy.yaml").read_text(encoding="utf-8"))
    runs = [_run("a", 8.6, 8.0, temp=tmp_path), _run("b", 8.2, 7.0, cert=False, temp=tmp_path)]
    out = tmp_path / "top.xlsx"
    write_workbook(runs, runs, policy, out, score_mode="single")
    wb = load_workbook(out)
    assert wb.sheetnames == ["总览", "硬性门槛", "内容质量", "数据口径", "原始证据"]
    overview = wb["总览"]
    headers = [c.value for c in overview[1]]
    assert "视频名称" in headers and "L3 均分" in headers and "定档" in headers
    gates = wb["硬性门槛"]
    assert gates.cell(row=2, column=1).value == "R01"
    assert gates.max_column == 9 + len(runs)
    quality = wb["内容质量"]
    assert any("前3秒钩子" == str(c.value) for c in quality[1])
    method = wb["数据口径"]
    assert any(r[0].value == "业务规则包" for r in method.iter_rows(min_row=2, max_col=1))
    evidence = wb["原始证据"]
    assert evidence.max_row >= 2 * 9 + 1  # 每 run 9 条关键证据


def test_discover_filters_published(tmp_path):
    import json as _json

    from scripts.export_top_videos import discover_runs

    # monkeypatch PROJECTS by pointing discover at tmp root with fake templates
    import scripts.export_top_videos as mod

    old = mod.PROJECTS
    mod.PROJECTS = tmp_path
    try:
        (tmp_path / "template-run-a").mkdir()
        (tmp_path / "template-run-a" / "checkpoint_publish.json").write_text(
            _json.dumps({"status": "completed"}), encoding="utf-8")
        (tmp_path / "template-run-b").mkdir()
        (tmp_path / "template-run-b" / "checkpoint_publish.json").write_text(
            _json.dumps({"status": "awaiting_human"}), encoding="utf-8")
        assert discover_runs() == ["template-run-a"]
    finally:
        mod.PROJECTS = old
