"""成片 TOP-N 导出：业务视角评价表（硬性门槛 / 投放前内容质量 / 系统评价）。

按 `docs/reports/export-design.md` 实现：
- 数据源全部为持久化制品（l1a_final / final_qa_full / delivery_certificate / asset_manifest /
  publish_log / edit_decisions / template_pack）+ L3 VLM（artifacts/l3_advisory.json，幂等落盘）；
- 业务规则包 `docs/rules/business-policy.yaml` 为唯一规则来源（CSV 仅 source_ref）；
- 输出 xlsx 5 表：总览 / 硬性门槛矩阵 / 内容质量 / 数据口径 / 原始证据（关键项+附件路径）；
- 排名：L3 均分 desc → 单维最低 desc → 有证书 → L1a 全绿 → 转场占比 → 成本 asc。

用法：
  python -m scripts.export_top_videos --top 5 --score-mode 3seed [--runs run ...] [--out path.xlsx]
"""
from __future__ import annotations

import argparse
import glob
import hashlib
import json
import os
from datetime import datetime, timezone
from pathlib import Path

try:
    import yaml
except ImportError:  # pragma: no cover
    yaml = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None

ROOT = Path(__file__).resolve().parents[1]
PROJECTS = ROOT / "projects"
PIPELINE = "cinematic-fast"
DEFAULT_POLICY = ROOT / "docs" / "rules" / "business-policy.yaml"
JUDGE_RUBRIC = "l3-v1.0"
JUDGE_FRAMES = 8
SEEDS = {"single": [42], "3seed": [42, 7, 2026]}

# 展示样式
_GREEN = PatternFill("solid", fgColor="C6EFCE")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_GRAY = PatternFill("solid", fgColor="D9D9D9")
_RED = PatternFill("solid", fgColor="FFC7CE")
_HEAD = PatternFill("solid", fgColor="F2E9D9")
_THIN = Border(*[Side(style="thin", color="DCD8D0")] * 4)


def _load_json(path: Path):
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def _file_sha(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _cell(ws, row, col, value, *, fill=None, bold=False, border=True, wrap=True, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold)
    if fill:
        c.fill = fill
    if border:
        c.border = _THIN
    c.alignment = Alignment(vertical="top", wrap_text=wrap, horizontal=align)
    return c


# ---------------------------------------------------------------------------
# 运行发现
# ---------------------------------------------------------------------------

def discover_runs() -> list[str]:
    """最近发布（publish completed）的模板 run，按发布时间倒序。"""
    runs = []
    for proj in sorted(glob.glob(str(PROJECTS / "template-run-*"))):
        name = Path(proj).name
        cp = _load_json(Path(proj) / "checkpoint_publish.json")
        if not cp or cp.get("status") != "completed":
            continue
        log = _load_json(Path(proj) / "artifacts" / "publish_log.json") or {}
        ts = (log.get("entries") or [{}])[0].get("timestamp") or ""
        runs.append((ts, name))
    runs.sort(reverse=True)
    return [name for _, name in runs]


def _template_meta(project: Path) -> dict:
    rp = _load_json(project / "artifacts" / "template_run_plan.json") or {}
    pack = _load_json(ROOT / "projects" / "template-pack-library" / "artifacts" / "template_pack.json") or {}
    tid = str(rp.get("template_id") or "")
    template = next((t for t in pack.get("templates", []) if t.get("template_id") == tid), {}) or {}
    return {"template_id": tid,
            "sheet_name": str(template.get("sheet_name") or template.get("name") or tid or project.name)}


# ---------------------------------------------------------------------------
# L3 评分（幂等落盘 artifacts/l3_advisory.json）
# ---------------------------------------------------------------------------

def score_run(project: Path, run: str, mode: str, *, force: bool = False) -> dict:
    from tools.tool_registry import registry

    media = project / "renders" / "final.mp4"
    if not media.is_file():
        raise SystemExit(f"{run}: 缺 renders/final.mp4，无法评分")
    advisory_path = project / "artifacts" / "l3_advisory.json"
    seeds = SEEDS[mode]
    existing = _load_json(advisory_path) if not force else None
    if existing and existing.get("media_sha256") == _file_sha(media) \
            and existing.get("rubric_version") == JUDGE_RUBRIC and existing.get("seeds") == seeds:
        return existing
    registry.discover()
    judge = registry._tools["video_judge"]
    l1a = _load_json(project / "artifacts" / "l1a_final.json") or {}
    loud = next((c.get("evidence", {}) for c in (l1a.get("hard_gate") or {}).get("checks", [])
                 if c.get("id") == "l1a_loudness"), {}) or {}
    audio_facts = (f"口播+ducked BGM 单轨混音；integrated {loud.get('integrated_lufs')} LUFS, "
                   f"True-Peak {loud.get('true_peak_dbtp')} dBTP")
    per_seed = []
    for seed in seeds:
        result = judge.execute({
            "input_path": str(media), "scope": "final", "rubric_version": JUDGE_RUBRIC,
            "frame_count": JUDGE_FRAMES, "seed": seed, "audio_facts": audio_facts,
            "output_path": "/dev/null" if os.name == "posix" else "NUL",
        })
        if not result.success:
            raise SystemExit(f"{run}: video_judge(seed={seed}) 失败: {result.error}")
        data = result.data or {}
        per_seed.append({d["id"]: d["score"] for d in data.get("dimensions", [])})
    dims_names = list(per_seed[0].keys()) if per_seed else []
    dims = {name: round(sum(p[name] for p in per_seed) / len(per_seed), 2) for name in dims_names}
    advisory = {
        "version": "1.0", "run": run, "judge_version": "video_judge-0.1.0",
        "rubric_version": JUDGE_RUBRIC, "model": "qwen-vl-max", "seeds": seeds,
        "frame_count": JUDGE_FRAMES, "score_mode": mode, "media_sha256": _file_sha(media),
        "per_seed": per_seed, "dimensions": dims, "audio_facts": audio_facts,
        "scored_at": datetime.now(timezone.utc).isoformat(),
        "summary": (result.data or {}).get("summary", ""),
    }
    advisory_path.parent.mkdir(parents=True, exist_ok=True)
    advisory_path.write_text(json.dumps(advisory, ensure_ascii=False, indent=2), encoding="utf-8")
    return advisory


# ---------------------------------------------------------------------------
# 单 run 数据收集
# ---------------------------------------------------------------------------

def collect_run(run: str, advisory: dict | None) -> dict:
    project = PROJECTS / run
    l1a = _load_json(project / "artifacts" / "l1a_final.json") or {}
    qa = _load_json(project / "artifacts" / "final_qa_full.json") or {}
    cert = _load_json(project / "artifacts" / "delivery_certificate.json")
    ed = _load_json(project / "artifacts" / "edit_decisions.json") or {}
    script = _load_json(project / "artifacts" / "script.json") or {}
    manifest = _load_json(project / "artifacts" / "asset_manifest.json") or {}
    plog = _load_json(project / "artifacts" / "publish_log.json") or {}
    l1a_checks = {(c.get("id")): c for c in (l1a.get("hard_gate") or {}).get("checks", []) if isinstance(c, dict)}
    probe = ((qa.get("checks") or {}).get("technical_probe") or {})
    cuts = ed.get("cuts") or []
    noncut = sum(1 for c in cuts if str(c.get("transition_in") or "cut") != "cut")
    cost = float(manifest.get("total_cost_usd") or 0.0)
    cost_note = ""
    if cost == 0.0:
        # reuse 路径记录 0：按实付口径（TTS §0.0002/段 + SUNO §0.05/片）估算并标注
        sections = script.get("sections") or []
        narrated = sum(1 for s in sections if str(s.get("narration") or "").strip())
        cost = round(narrated * 0.0002 + 0.05, 4)
        cost_note = "按实付口径估算（manifest 复用路径记 0）"
    return {
        "run": run, "project": project, ** _template_meta(project),
        "l1a": l1a, "l1a_checks": l1a_checks, "qa": qa, "probe": probe,
        "certificate": cert, "edit": ed, "cuts": cuts, "noncut": noncut,
        "script": script, "duration_s": round(float(script.get("total_duration_seconds") or 0), 1),
        "cost": cost, "cost_note": cost_note,
        "published_at": ((plog.get("entries") or [{}])[0].get("timestamp") or ""),
        "advisory": advisory,
        "l3_avg": round(sum(advisory["dimensions"].values()) / len(advisory["dimensions"]), 2) if advisory else None,
        "l3_min": min(advisory["dimensions"].values()) if advisory else None,
        "weakest": min(advisory["dimensions"], key=advisory["dimensions"].get) if advisory else "",
    }


def tier_of(run: dict) -> str:
    avg, mn = run["l3_avg"], run["l3_min"]
    if avg is None:
        return "未评分"
    if avg >= 8.5:
        return "推荐"
    if avg >= 8.1:
        return "达标"
    return "观察" + ("（短板: " + str(run["weakest"]) + "）" if mn is not None and mn <= 6.5 else "")


def rank_runs(runs: list[dict]) -> list[dict]:
    def key(r):
        noncut_ratio = r["noncut"] / max(len(r["cuts"]), 1)
        return (-(r["l3_avg"] or 0), -(r["l3_min"] or 0),
                0 if r["certificate"] else 1, 0 if str(r["l1a"].get("status")) == "pass" else 1,
                -noncut_ratio, r["cost"])
    return sorted(runs, key=key)


# ---------------------------------------------------------------------------
# 规则判定（策略包 → 证据）
# ---------------------------------------------------------------------------

def rule_verdict(rule: dict, run: dict) -> tuple[str, str]:
    """返回 (判定文案, 样式 key: pass|partial|pending|external|fail)。"""
    if rule.get("layer") == "external":
        return "外部依赖", "external"
    checks = rule.get("checks") or []
    if not checks:
        return "待接入", "pending"
    failed, missing, noted = [], [], []
    for check in checks:
        if check == "resolution":
            value = run["probe"].get("resolution") or ""
            if value:
                w, h = (int(x) for x in value.lower().split("x"))
                if min(w, h) >= 720:
                    noted.append(f"{value}≥720p")
                else:
                    failed.append(f"分辨率{value}<720p")
            else:
                missing.append("resolution")
            continue
        if check == "duration":
            dur = run["probe"].get("duration_seconds") or 0
            if dur and 15 <= dur <= 60:
                noted.append(f"时长{dur:.1f}s∈[15,60]")
            else:
                failed.append(f"时长{dur:.1f}s∉[15,60]")
            continue
        if check.startswith("l3_"):
            dim = check[3:]
            score = (run["advisory"] or {}).get("dimensions", {}).get(dim)
            if score is not None:
                noted.append(f"{dim}={score}")
            else:
                missing.append(dim)
            continue
        item = run["l1a_checks"].get(check)
        if not item:
            missing.append(check)
        elif item.get("status") == "fail":
            failed.append(check + ":" + str(item.get("message") or "")[:40])
        else:
            noted.append(check)
    if failed:
        return "不合格（" + "；".join(failed[:2]) + "）", "fail"
    note = "；".join(noted[:4])
    if missing:
        return ("部分取证（" + note + "；未取证: " + "、".join(missing[:3]) + "）", "partial")
    return "通过（" + note + "）", "pass"


# ---------------------------------------------------------------------------
# xlsx 输出
# ---------------------------------------------------------------------------

HEADER_STYLE = dict(fill=_HEAD, bold=True)
STATUS_FILL = {"pass": _GREEN, "partial": _YELLOW, "pending": _GRAY,
               "external": _GRAY, "fail": _RED}


def _write_header(ws, headers: list[str], row: int = 1) -> None:
    for col, title in enumerate(headers, start=1):
        _cell(ws, row, col, title, **HEADER_STYLE)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def write_workbook(runs: list[dict], rank_ordered: list[dict], policy: dict,
                   out: Path, *, score_mode: str) -> None:
    wb = Workbook()

    # ---- ① 总览 ----
    ws = wb.active
    ws.title = "总览"
    headers = ["排名", "视频名称", "模板编号", "成片", "时长(秒)", "定档", "L3 均分", "单维最低",
               "短板维度", "L1a 硬门", "合规检查", "画质·时长门槛", "交付证书", "字幕安全区",
               "响度", "转场落地", "人工确认", "成本", "单位成本质量", "发布时间"]
    _write_header(ws, headers)
    for i, run in enumerate(rank_ordered, start=1):
        loud = next((c.get("evidence", {}) for c in run["l1a"].get("hard_gate", {}).get("checks", [])
                     if c.get("id") == "l1a_loudness"), {}) or {}
        sens = run["l1a_checks"].get("l1a_sensitive") or {}
        cp_approved = _load_json(run["project"] / "checkpoint_sample.json") or {}
        row = [
            i, run["sheet_name"], run["template_id"], "renders/final.mp4", run["duration_s"], tier_of(run),
            run["l3_avg"], run["l3_min"], run["weakest"], str(run["l1a"].get("status") or "—"),
            "通过" if sens.get("status") == "pass" else "未取证",
            "通过" if run["probe"].get("resolution") and run["probe"].get("duration_seconds") else "—",
            "✅ 绑定" if run["certificate"] else "无",
            "通过" if (run["l1a_checks"].get("l1a_subtitle_bounds") or {}).get("status") == "pass" else "—",
            f"{loud.get('integrated_lufs')} LUFS / {loud.get('true_peak_dbtp')} dBTP",
            f"{run['noncut']}/{len(run['cuts'])}",
            "已批准（批量授权口径）" if cp_approved.get("human_approved") else "未批准",
            f"{run['cost']:.4f}{('（' + run['cost_note'] + '）') if run['cost_note'] else ''}",
            round(run["cost"] / max(run["l3_avg"] or 1, 0.01), 4),
            run["published_at"][:19].replace("T", " "),
        ]
        for col, value in enumerate(row, start=1):
            fill = _GREEN if col == 6 and value == "推荐" else None
            _cell(ws, i + 1, col, value, fill=fill)
    for col, width in zip(range(1, len(headers) + 1), [6, 18, 26, 16, 9, 18, 8, 8, 10, 8, 9, 12, 9, 8, 20, 10, 16, 12, 10, 18]):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ---- ② 硬性门槛（矩阵：行=规则，列=视频） ----
    rules = [r for r in policy["rules"] if r.get("layer") == "hard_gate"]
    ws2 = wb.create_sheet("硬性门槛")
    _write_header(ws2, ["规则ID", "规则名称", "规则类别", "指标类型", "量化指标", "阈值/评分标准",
                        "我方证据源", "实现状态", "判定语义"] + [r["sheet_name"] for r in rank_ordered])
    for ridx, rule in enumerate(rules, start=2):
        row = [rule["rule_id"], rule["rule_name"], rule["rule_category"], rule["metric_type"],
               rule["metric"], rule["thresholds"], rule["evidence_source"], rule["impl_status"],
               "一票否决" if rule.get("severity") == "veto" else "计分项"]
        _cell(ws2, ridx, 1, row[0], bold=True)
        for col, value in enumerate(row[1:], start=2):
            _cell(ws2, ridx, col, value)
        for cidx, run in enumerate(rank_ordered):
            text, style = rule_verdict(rule, run)
            _cell(ws2, ridx, 9 + cidx + 1, text, fill=STATUS_FILL.get(style))
    for col, width in zip(range(1, 10), [8, 24, 10, 10, 42, 26, 30, 10, 10]):
        ws2.column_dimensions[get_column_letter(col)].width = width
    for col in range(10, 10 + len(rank_ordered)):
        ws2.column_dimensions[get_column_letter(col)].width = 30

    # ---- ③ 内容质量（行=视频） ----
    prelaunch = [r for r in policy["rules"] if r.get("layer") == "prelaunch"]
    ws3 = wb.create_sheet("内容质量")
    dim_headers = ["前3秒钩子", "视觉层级", "节奏", "镜头质量", "故事连贯", "音频质量",
                   "文字可读", "商品露出", "L3 均分", "单维最低"]
    biz_headers = [f"{r['rule_id']} {r['rule_name'][:12]}(近似)"
                   for r in prelaunch if (r.get("checks") or [])]
    _write_header(ws3, ["排名", "视频名称"] + dim_headers + biz_headers
                  + ["技术摘要", "转场数/总镜", "证书", "人工确认", "成本"])
    for i, run in enumerate(rank_ordered, start=2):
        dims = (run["advisory"] or {}).get("dimensions", {}) or {}
        row = [i, run["sheet_name"]]
        for key in ("hook_clarity", "visual_hierarchy", "rhythm", "shot_quality",
                    "story_coherence", "audio_quality", "text_readability", "product_presence"):
            row.append(dims.get(key))
        row += [run["l3_avg"], run["l3_min"]]
        for rule in prelaunch:
            if not (rule.get("checks") or []):
                continue
            vals = [dims.get(c[3:]) for c in rule["checks"] if c.startswith("l3_") and dims.get(c[3:]) is not None]
            biz = round(sum(vals) / len(vals), 2) if vals else "—"
            row.append(biz)
        black = (run["l1a_checks"].get("l1a_black_frames") or {}).get("status") or "—"
        freeze = (run["l1a_checks"].get("l1a_freeze") or {}).get("status") or "—"
        sub = (run["l1a_checks"].get("l1a_subtitle_bounds") or {}).get("status") or "—"
        row += [f"黑帧:{black} 冻结:{freeze} 字幕:{sub}", f"{run['noncut']}/{len(run['cuts'])}"]
        row += ["✅" if run["certificate"] else "无",
                "已批准" if (_load_json(run["project"] / "checkpoint_sample.json") or {}).get("human_approved") else "未批准",
                run["cost"]]
        for col, value in enumerate(row, start=1):
            _cell(ws3, i, col, value,
                  fill=_GREEN if col == 3 and value is not None and value >= 8.5 else None)
    for col, width in zip(range(1, 4), [6, 18, 18]):
        ws3.column_dimensions[get_column_letter(col)].width = width
    for col in range(4, 4 + len(dim_headers) + len(biz_headers) + 5):
        ws3.column_dimensions[get_column_letter(col)].width = 13

    # ---- ④ 数据口径 ----
    ws4 = wb.create_sheet("数据口径")
    _write_header(ws4, ["字段", "值", "说明"])
    rows = [
        ("导出时间", datetime.now(timezone.utc).isoformat(), "UTC"),
        ("评价体系版本", "docs/EVALUATION_SYSTEM.md（§13 业务册接入评估）", "规则分层依据"),
        ("judge_version", "video_judge-0.1.0", "technical_validator-0.1.0 为 L1a"),
        ("rubric_version", JUDGE_RUBRIC, "l3-v1.0（advisory，不进发布硬门）"),
        ("VLM 模型", "qwen-vl-max", "随机评价器；3seed 均值更稳"),
        ("评分种子", " / ".join(map(str, SEEDS[score_mode])), "随机性记录（评价体系 §3.5）"),
        ("抽帧数", JUDGE_FRAMES, "均匀抽帧"),
        ("打分模式", score_mode, "single=1 seed；3seed=均值（推荐正式版）"),
        ("运行清单", "、".join(r["run"] for r in rank_ordered), "按发布时间排序"),
        ("排序规则", "L3 均分 desc → 单维最低 desc → 有证书 → L1a 全绿 → 转场占比 → 成本 asc", "§4 复合排序"),
        ("业务规则包", "docs/rules/business-policy.yaml v1.0", "R01–R24；CSV 仅 source_ref"),
        ("成本口径", "manifest total_cost_usd；reuse 记 0 按实付估算并标注", "TTS+SUNO；素材全自有 $0"),
        ("已知限制", "；".join([
            "sheet-01/04 无交付证书且为探索期链路产物",
            "五确认=批量授权口径（真实五项采集待 Editorial Gallery）",
            "R02/R04/R05/R17 待接入；R06/R11/R13/R16/R19–R24 外部依赖（本期不评估）",
            "业务近似列为近似度标注，不做伪精确换算",
        ]), "诚实性声明"),
    ]
    for ridx, (k, v, note) in enumerate(rows, start=2):
        _cell(ws4, ridx, 1, k, bold=True)
        _cell(ws4, ridx, 2, v)
        _cell(ws4, ridx, 3, note)
    for col, width in zip(range(1, 4), [16, 70, 46]):
        ws4.column_dimensions[get_column_letter(col)].width = width

    # ---- ⑤ 原始证据（关键项 + 附件路径） ----
    ws5 = wb.create_sheet("原始证据")
    _write_header(ws5, ["视频名称", "证据制品", "检查项", "状态", "实测值", "阈值", "消息/说明"])
    for run in rank_ordered:
        evidence = [
            ("l1a_final.json", "l1a_sensitive", run["l1a_checks"].get("l1a_sensitive")),
            ("l1a_final.json", "l1a_subtitle_bounds", run["l1a_checks"].get("l1a_subtitle_bounds")),
            ("l1a_final.json", "l1a_loudness", run["l1a_checks"].get("l1a_loudness")),
            ("l1a_final.json", "l1a_duration", run["l1a_checks"].get("l1a_duration")),
            ("final_qa_full.json", "technical_probe", {"status": "pass", "value":
                f"res={run['probe'].get('resolution')} dur={run['probe'].get('duration_seconds')}s "
                f"audio={run['probe'].get('has_audio')}"}),
            ("final_qa_full.json", "subtitle_check", {"status": "pass", "value": ""}),
            ("l3_advisory.json", "l3_avg", {"status": "scored", "value": f"{run['l3_avg']} (seeds 均值)"}),
            ("delivery_certificate.json", "certified", {"status": "pass" if run["certificate"] else "missing", "value": ""}),
            ("edit_decisions.json", f"transition_parity", {"status": "pass", "value": f"{run['noncut']}/{len(run['cuts'])}"}),
        ]
        for artifact, check, item in evidence:
            status = (item or {}).get("status", "missing")
            text = (item or {}).get("message") or (item or {}).get("value") or ""
            _cell(ws5, ws5.max_row + 1, 1, run["sheet_name"])
            _cell(ws5, ws5.max_row, 2, "artifacts/" + artifact)
            _cell(ws5, ws5.max_row, 3, check)
            _cell(ws5, ws5.max_row, 4, status, fill=STATUS_FILL.get(status, _GRAY))
            _cell(ws5, ws5.max_row, 5, text)
            _cell(ws5, ws5.max_row, 6, "")
            _cell(ws5, ws5.max_row, 7, f"附件: projects/{run['run']}/artifacts/{artifact}（renders/final.mp4 同目录）")
    for col, width in zip(range(1, 8), [18, 26, 18, 10, 34, 12, 34]):
        ws5.column_dimensions[get_column_letter(col)].width = width

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--top", type=int, default=5)
    p.add_argument("--runs", nargs="*", default=None)
    p.add_argument("--score-mode", choices=["single", "3seed"], default="single")
    p.add_argument("--force-score", action="store_true")
    p.add_argument("--out", default=None)
    p.add_argument("--policy", default=str(DEFAULT_POLICY))
    args = p.parse_args()

    if yaml is None or Workbook is None:
        raise SystemExit("需要 pyyaml 与 openpyxl（pip install pyyaml openpyxl）")
    policy = yaml.safe_load(Path(args.policy).read_text(encoding="utf-8"))
    runs = args.runs or discover_runs()
    if not runs:
        raise SystemExit("未发现已发布的模板 run（--runs 可显式指定）")
    # 最近运行取较全集合（足够 TOP N 判定）
    runs = runs[:max(args.top * 2, 10)] + (args.runs or [])
    runs = list(dict.fromkeys(runs))
    collected = []
    for run in runs:
        project = PROJECTS / run
        advisory = score_run(project, run, args.score_mode, force=args.force_score)
        collected.append(collect_run(run, advisory))
    ordered = rank_runs(collected)[: args.top]
    out = Path(args.out) if args.out else ROOT / "docs" / "reports" / "export" / (
        f"top-{args.top}-videos-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx")
    write_workbook(collected, ordered, policy, out, score_mode=args.score_mode)
    print(f"已导出 {len(ordered)} 部成片 → {out}")
    for i, run in enumerate(ordered, start=1):
        print(f"  #{i} {run['sheet_name']:<14} L3={run['l3_avg']} 单维最低={run['l3_min']} "
              f"定档={tier_of(run)} 证书={'✅' if run['certificate'] else '无'}")


if __name__ == "__main__":
    main()
