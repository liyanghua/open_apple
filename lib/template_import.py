"""Template pack import from the 43-sheet xlsx (参考模板库)。

把 `docs/insight_source/视频分镜拆解_2026-08-15.xlsx`（43 sheet = 43 条不同参考视频，
每条 14 列人工逐镜拆解）解析成本项目可复用的只读 ``template_pack`` 制品。

字段映射（必须按人工列，不猜测）：
- H 列「花字」(col 8) → ``overlay_text``（文字内容）；
- I 列「特效」(col 9) → ``caption_treatment``（处理方式：淡入/字幕/字幕动画/淡出…）。

导入是**人工数据的事实锁定**，不是让 video_analyzer 重新猜测人工列。重复导入同一
source hash 结果一致；source hash 或 parser/taxonomy version 变化产生新 artifact hash。
"""
from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

# 14 列（索引从 1 计，含表头）。用 0 基索引映射到行 value 列表。
COL_MAP = {
    "ordinal": 0,            # 序号
    "shot_size": 1,          # 景别
    "camera_movement": 2,    # 镜头
    "camera_angle": 3,       # 拍摄方法
    "duration": 4,           # 时长
    "visual_content": 5,     # 画面内容
    "dialogue": 6,           # 台词
    "overlay_text": 7,       # 花字 → overlay_text
    "caption_treatment": 8,  # 特效 → caption_treatment
    "notes": 9,              # 备注
    "visual_reference": 10,  # 画面参考
    "scene": 11,             # 场景
    "audio_type": 12,        # 音频类型
    "bgm": 13,               # BGM
}
COL_COUNT = 14

# 特效列 → caption_treatment 枚举。
CAPTION_TREATMENT_MAP = {
    "淡入": "fade_in",
    "淡出": "fade_out",
    "字幕动画": "animated",
    "字幕动效": "animated",
    "动效": "animated",
    "字幕": "subtitle",
}

# 已知商品名 → ascii slug（template_id 用；未知商品走降级 slug）。
_PRODUCT_SLUG_MAP = {
    "AKS桌垫": "aks-zhuodian",
    "桌垫": "zhuodian",
    "岩板桌架": "yanban-zhuojia",
}

PARSER_VERSION = "xlsx-template-import@1"
TAXONOMY_VERSION = "template-pack@1"


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _parse_interval(duration: str | None) -> dict[str, float | None]:
    t = str(duration or "").strip()
    m = re.match(r"^\s*([\d.]+)\s*-\s*([\d.]+)\s*s?", t)
    if not m:
        return {"start_s": None, "end_s": None, "duration_s": None}
    start, end = float(m.group(1)), float(m.group(2))
    return {"start_s": start, "end_s": end, "duration_s": max(0.0, end - start)}


def _normalize_caption_treatment(raw: str | None, overlay_text: str | None) -> tuple[str, bool]:
    """I 特效列 → caption_treatment；返回 (treatment, warning)。"""
    s = str(raw or "").strip()
    if s:
        mapped = CAPTION_TREATMENT_MAP.get(s)
        if mapped:
            return mapped, False
        return "unknown", True  # 未知特效 → 人工复核
    if str(overlay_text or "").strip():
        return "static", False  # 有花字但未记录动画
    return "none", False  # 该镜无花字处理


def _make_template_id(sheet_name: str, idx: int) -> str:
    """稳定 template_id：sheet 序号 + sheet 名中的视频编号 + 商品 slug。

    sheet 名如 `视频1_AKS桌垫` / `视频48_岩板桌架`；视频编号不连续（缺 11/16/…），
    所以不能用连续序号推断，必须从 sheet 名解析。
    """
    m = re.match(r"视频(\d+)_(.+)", sheet_name or "")
    vid = m.group(1) if m else str(idx)
    product = m.group(2) if m else "unknown"
    slug = _PRODUCT_SLUG_MAP.get(product) or _ascii_slug(product)
    return f"sheet-{idx:02d}-video{vid}-{slug}"


def _ascii_slug(text: str) -> str:
    """未知商品降级 slug：仅保留 ascii 数字/字母，转小写，空则 'unknown'。"""
    cleaned = re.sub(r"[^A-Za-z0-9]+", "-", text).strip("-").lower()
    return cleaned or "unknown"


def _normalize_row(row: tuple | None) -> list:
    if not row:
        return [None] * COL_COUNT
    values = list(row)
    if len(values) < COL_COUNT:
        values += [None] * (COL_COUNT - len(values))
    return values[:COL_COUNT]


def _extract_visual_references(ws, template_id: str, evidence_dir: Path | None) -> dict[int, dict[str, Any]]:
    """从 sheet 的 K 列（画面参考）内嵌图片导出，返回 {row(1-based): visual_reference_ref}。

    ``values_only=True`` 读不到图片；openpyxl 的 ``ws._images`` 保留 anchor + 字节。
    导出到 evidence_dir/template_id/，按源单元格命名 + 记录 sha256。
    """
    out: dict[int, dict[str, Any]] = {}
    if evidence_dir is None:
        return out
    for img in getattr(ws, "_images", []):
        anchor = getattr(img, "anchor", None)
        if anchor is None:
            continue
        try:
            col = int(anchor._from.col)  # 0-based
            excel_row = int(anchor._from.row) + 1  # 1-based Excel 行（含表头行1）
        except (AttributeError, TypeError):
            continue
        if col != COL_MAP["visual_reference"]:
            continue  # 只认 K 列（画面参考）
        try:
            data = img._data()
        except (AttributeError, TypeError):
            continue
        if not data:
            continue
        # data ordinal = excel_row - 1（表头占行1）；slot loop 的 r_i 从 1 起（min_row=2）
        data_ordinal = excel_row - 1
        if data_ordinal < 1:
            continue
        ext = ({'png': 'png', 'jpeg': 'jpg', 'gif': 'gif'}.get(str(getattr(img, 'format', 'png') or 'png').lower(), 'png'))
        cell = f"K{excel_row}"
        fname = f"slot-{data_ordinal:03d}-{cell}.{ext}"
        fpath = evidence_dir / template_id / fname
        fpath.parent.mkdir(parents=True, exist_ok=True)
        fpath.write_bytes(data)
        out[data_ordinal] = {
            "evidence_ref": fpath.relative_to(evidence_dir).as_posix(),
            "source_cell": cell,
            "image_sha256": hashlib.sha256(data).hexdigest(),
        }
    return out


def _cell(values: list, key: str) -> Any:
    idx = COL_MAP[key]
    return values[idx] if idx < len(values) else None


def build_template_pack(
    xlsx_path: str | Path,
    *,
    project_id: str = "template-pack",
    parser_version: str = PARSER_VERSION,
    evidence_dir: str | Path | None = None,
) -> dict[str, Any]:
    """解析 xlsx → template_pack dict（含 sha256 + warnings）。"""
    import openpyxl

    path = Path(xlsx_path)
    if not path.is_file():
        raise FileNotFoundError(f"xlsx not found: {path}")
    source = path.read_bytes()
    source_sha = _sha256(source)
    # 可复现身份：created_at 用源文件 mtime（同一次文件内容稳定），path 归一化为仓库相对路径。
    # 否则同一文件连续导入（now() 不同/绝对 vs 相对路径不同）会得到不同的 artifact/semantic hash。
    root = Path(__file__).resolve().parents[1]
    try:
        rel_path = path.resolve().relative_to(root).as_posix()
    except ValueError:
        rel_path = path.name
    created_at = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()

    wb = openpyxl.load_workbook(path, data_only=True)
    templates: list[dict[str, Any]] = []
    warnings: list[str] = []
    ev_dir = Path(evidence_dir) if evidence_dir else None

    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        template_id = _make_template_id(sheet_name, idx)
        refs_by_row = _extract_visual_references(ws, template_id, ev_dir)
        slots: list[dict[str, Any]] = []
        for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            if row is None or all(v is None for v in row):
                continue
            values = _normalize_row(row)
            overlay = _cell(values, "overlay_text")
            treatment, warn = _normalize_caption_treatment(
                _cell(values, "caption_treatment"), overlay
            )
            if warn:
                warnings.append(f"{sheet_name} row{r_i}: 未知特效 {_cell(values, 'caption_treatment')!r}")
            interval = _parse_interval(_cell(values, "duration"))
            slot = {
                "slot_id": f"{template_id}-slot-{r_i:03d}",
                "ordinal": r_i,
                "duration_s": interval["duration_s"],
                "shot_language": {
                    "shot_size": _cell(values, "shot_size"),
                    "camera_movement": _cell(values, "camera_movement"),
                    "camera_angle": _cell(values, "camera_angle"),
                },
                "visual_content": _cell(values, "visual_content"),
                "overlay_text": overlay,
                "caption_treatment": treatment,
                "effect_treatment": _cell(values, "caption_treatment"),
                "audio_layers": [str(_cell(values, "audio_type"))] if _cell(values, "audio_type") else [],
                "music_profile": _cell(values, "bgm"),
                "scene": _cell(values, "scene"),
                "dialogue": _cell(values, "dialogue"),
            }
            if refs_by_row.get(r_i):
                slot["visual_reference_ref"] = refs_by_row[r_i]
            slots.append(slot)
        if slots:
            templates.append({
                "template_id": template_id,
                "sheet_name": sheet_name,
                "archetype": None,
                "slots": slots,
            })

    pack: dict[str, Any] = {
        "version": "1.0",
        "project_id": project_id,
        "created_at": created_at,
        "taxonomy_version": TAXONOMY_VERSION,
        "source_document": {
            "path": rel_path,
            "sha256": source_sha,
            "parser_version": parser_version,
        },
        "templates": templates,
        "normalization_warnings": warnings,
    }
    return pack
